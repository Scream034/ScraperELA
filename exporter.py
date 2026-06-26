"""
ScraperELA · exporter.py
=========================
Экспорт из единой SQLite в профессиональный XLSX.

Особенности форматирования:

  Статус ФНС:
    Поле всегда заполнено. Возможные значения:
      • Обычный статус (ACTIVE, LIQUIDATED, …) — цветная заливка + метка.
      • «LIQUIDATED|2021-06-17» — статус с встроенной датой через «|».
      • «NO_INN» — нет ИНН, проверка по ЕГРЮЛ невозможна (серый).
      • «PENDING» — ИНН есть, но проверка ещё не выполнялась (голубой).

  КПП:
    Несколько КПП отображаются в столбик (char(10) разделитель в SQL).

  Дата ликвидации:
    Встроена в ячейку «Статус ФНС» через перенос строки:
    «✗ Ликвидировано↵2021-06-17». Отдельной колонки нет.
"""

from __future__ import annotations

import datetime
import logging
import math
import sqlite3
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import config as cfg

logger = logging.getLogger("Exporter")


# ---------------------------------------------------------------------------
# Цвета и статусы
# ---------------------------------------------------------------------------


class _C:
    HEADER_BG = "1F4E79"
    HEADER_FG = "FFFFFF"
    ROW_EVEN = "F2F7FF"
    ROW_ODD = "FFFFFF"
    STATUS_ACTIVE = "C6EFCE"
    STATUS_LIQUIDATED = "FFC7CE"
    STATUS_LIQUIDATING = "FFEB9C"
    STATUS_REORGANIZING = "FFEB9C"
    STATUS_BANKRUPT = "FFC7CE"
    STATUS_UNKNOWN = "D9D9D9"
    STATUS_NO_INN = "EFEFEF"
    STATUS_PENDING = "DDEEFF"
    INN_BG = "EBF3FB"


_STATUS_COLOR: dict[str, str] = {
    "ACTIVE": _C.STATUS_ACTIVE,
    "LIQUIDATED": _C.STATUS_LIQUIDATED,
    "LIQUIDATING": _C.STATUS_LIQUIDATING,
    "REORGANIZING": _C.STATUS_REORGANIZING,
    "BANKRUPT": _C.STATUS_BANKRUPT,
    "UNKNOWN": _C.STATUS_UNKNOWN,
    "NO_INN": _C.STATUS_NO_INN,
    "PENDING": _C.STATUS_PENDING,
}

_STATUS_LABEL: dict[str, str] = {
    "ACTIVE": "✓ Действующее",
    "LIQUIDATED": "✗ Ликвидировано",
    "LIQUIDATING": "⚠ Ликвидация",
    "REORGANIZING": "⚠ Реорганизация",
    "BANKRUPT": "✗ Банкрот",
    "UNKNOWN": "? Не определён",
    "NO_INN": "— Нет ИНН",
    "PENDING": "⏳ Ожидает проверки",
}

# Допустимые поля сортировки → SQL-выражение
_SORT_FIELDS: dict[str, str] = {
    "name": "c.name",
    "inn": "c.inn",
    "address": "c.address",
    "director": "c.director",
    "scrape_date": "c.scrape_date",
    "first_seen_at": "c.first_seen_at",
    "last_seen_at": "c.last_seen_at",
    "status_official": "c.status_official",
    "source_count": (
        "(SELECT COUNT(*) FROM company_sources cs3"
        " WHERE cs3.company_id = c.company_id)"
    ),
}


# ---------------------------------------------------------------------------
# Конфигурация колонок
# ---------------------------------------------------------------------------

_MULTILINE: frozenset[str] = frozenset(
    {
        "Телефоны",
        "Email",
        "Источники",
        "КПП",
        "Адрес (агрегатор)",
        "Офиц. адрес (ЕГРЮЛ)",
        "Название",
        "Юр. название",
        "Офиц. наименование (ЕГРЮЛ)",
        "Сайт",
        "Статус ФНС",
    }
)

_CENTER: frozenset[str] = frozenset(
    {
        "ИНН",
        "ОГРН",
        "Дата регистрации",
        "Дата проверки ФНС",
        "Дата парсинга",
    }
)

_DATE_COLS: frozenset[str] = frozenset(
    {
        "Дата регистрации",
        "Дата проверки ФНС",
        "Дата парсинга",
    }
)

_TEXT_NUM: frozenset[str] = frozenset({"ИНН", "ОГРН", "КПП"})

_MIN_W = 12
_MAX_W = 55
_LINE_H = 15.5

# Разделитель между кодом статуса и датой ликвидации в SQL-результате.
_STATUS_DATE_SEP = "|"


# ---------------------------------------------------------------------------
# Динамический построитель SQL
# ---------------------------------------------------------------------------


def _build_export_sql(export_cfg: dict[str, Any]) -> tuple[str, list[Any]]:
    """Строит SQL-запрос экспорта с фильтрами и сортировкой.

    КПП выводится через ``char(10)`` (перенос строки) вместо ``', '``.

    Поле «Статус ФНС» формируется по правилам:
      • Есть статус + дата ликвидации → ``"LIQUIDATED|2021-06-17"``
      • Есть статус без даты          → ``"ACTIVE"``
      • Нет статуса, нет ИНН          → ``"NO_INN"``
      • Нет статуса, есть ИНН         → ``"PENDING"``

    Args:
        export_cfg: ``config.EXPORT_CONFIG``.

    Returns:
        Кортеж ``(sql_string, params_list)``.
    """
    where_parts: list[str] = []
    params: list[Any] = []

    if city := export_cfg.get("filter_city"):
        where_parts.append("c.address LIKE ? COLLATE NOCASE")
        params.append(f"%{city}%")

    if site_key := export_cfg.get("filter_site_key"):
        where_parts.append("""
            EXISTS (
                SELECT 1 FROM company_sources cs_f
                 WHERE cs_f.company_id = c.company_id
                   AND cs_f.site_key   = ?
            )
        """)
        params.append(site_key)

    if status_official := export_cfg.get("filter_status_official"):
        where_parts.append("c.status_official = ?")
        params.append(status_official)

    if export_cfg.get("filter_has_inn"):
        where_parts.append("c.inn IS NOT NULL")

    if export_cfg.get("filter_has_phone"):
        where_parts.append("""
            EXISTS (
                SELECT 1 FROM company_contacts cc_f
                 WHERE cc_f.company_id   = c.company_id
                   AND cc_f.contact_type = 'phone'
            )
        """)

    where_sql = f"WHERE {' AND '.join(where_parts)}" if where_parts else ""

    sort_parts: list[str] = []
    for field, direction in export_cfg.get("sort_by", [("name", "ASC")]):
        if sql_expr := _SORT_FIELDS.get(field):
            dir_sql = "DESC" if str(direction).upper() == "DESC" else "ASC"
            sort_parts.append(f"{sql_expr} {dir_sql} NULLS LAST")
        else:
            logger.warning(f"Неизвестное поле сортировки '{field}'.")
    order_sql = (
        f"ORDER BY {', '.join(sort_parts)}"
        if sort_parts
        else "ORDER BY c.name NULLS LAST"
    )

    limit = int(export_cfg.get("limit", 0))
    limit_sql = f"LIMIT {limit}" if limit > 0 else ""

    sql = f"""
        SELECT
            c.name                                                  AS "Название",
            CASE
                WHEN c.status_official IS NOT NULL
                     AND c.liquidation_date IS NOT NULL
                    THEN c.status_official
                         || '{_STATUS_DATE_SEP}'
                         || c.liquidation_date
                WHEN c.status_official IS NOT NULL
                    THEN c.status_official
                WHEN c.inn IS NULL
                    THEN 'NO_INN'
                ELSE 'PENDING'
            END                                                     AS "Статус ФНС",
            c.address                                               AS "Адрес (агрегатор)",
            c.address_official                                      AS "Офиц. адрес (ЕГРЮЛ)",
            (
                SELECT GROUP_CONCAT(cc.value, char(10))
                  FROM company_contacts cc
                 WHERE cc.company_id   = c.company_id
                   AND cc.contact_type = 'phone'
            )                                                       AS "Телефоны",
            (
                SELECT GROUP_CONCAT(cc.value, char(10))
                  FROM company_contacts cc
                 WHERE cc.company_id   = c.company_id
                   AND cc.contact_type = 'email'
            )                                                       AS "Email",
            c.director                                              AS "Директор (агрегатор)",
            c.director_official                                     AS "Офиц. руководитель (ЕГРЮЛ)",
            c.inn                                                   AS "ИНН",
            c.ogrn                                                  AS "ОГРН",
            (
                SELECT GROUP_CONCAT(ck.kpp, char(10))
                  FROM company_kpp ck
                 WHERE ck.company_id = c.company_id
            )                                                       AS "КПП",
            c.legal_name                                            AS "Юр. название",
            c.legal_name_official                                   AS "Офиц. наименование (ЕГРЮЛ)",
            c.registration_date                                     AS "Дата регистрации",
            c.official_verified_at                                  AS "Дата проверки ФНС",
            c.scrape_date                                           AS "Дата парсинга",
            c.website                                               AS "Сайт",
            (
                SELECT GROUP_CONCAT(cs.source_url, char(10))
                  FROM company_sources cs
                 WHERE cs.company_id = c.company_id
                 ORDER BY cs.first_seen_at
            )                                                       AS "Источники"
        FROM companies c
        {where_sql}
        {order_sql}
        {limit_sql}
    """

    return sql, params


def _describe_filters(export_cfg: dict[str, Any]) -> str:
    """Формирует читаемую строку активных фильтров для лога."""
    parts: list[str] = []
    if v := export_cfg.get("filter_city"):
        parts.append(f"город='{v}'")
    if v := export_cfg.get("filter_site_key"):
        parts.append(f"источник='{v}'")
    if v := export_cfg.get("filter_status_official"):
        parts.append(f"статус_фнс='{v}'")
    if export_cfg.get("filter_has_inn"):
        parts.append("только_с_инн")
    if export_cfg.get("filter_has_phone"):
        parts.append("только_с_телефоном")
    if (v := export_cfg.get("limit")) and int(v) > 0:
        parts.append(f"лимит={v}")
    if sort_by := export_cfg.get("sort_by", []):
        parts.append(f"сортировка=[{', '.join(f'{f} {d}' for f, d in sort_by)}]")
    return " | ".join(parts) if parts else "без фильтров"


# ---------------------------------------------------------------------------
# Вспомогательные функции
# ---------------------------------------------------------------------------


def _make_border() -> Border:
    s = Side(border_style="thin", color="C8C8C8")
    return Border(left=s, right=s, top=s, bottom=s)


def _parse_date(v: Any) -> datetime.date | None:
    if not isinstance(v, str) or not v:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.datetime.strptime(v, fmt).date()
        except ValueError:
            continue
    return None


def _row_fill(row_idx: int) -> PatternFill:
    c = _C.ROW_EVEN if row_idx % 2 == 0 else _C.ROW_ODD
    return PatternFill(fill_type="solid", fgColor=c)


def _estimate_lines(text: str, col_width: float) -> int:
    """Оценивает число визуальных строк с учётом \\n и автопереноса."""
    if not text:
        return 1
    max_chars = max(1, int(col_width * 1.2))
    total = 0
    for line in text.split("\n"):
        total += max(1, math.ceil(len(line) / max_chars)) if line else 1
    return max(1, total)


def _extract_status_code(raw: str | None) -> str | None:
    """Извлекает код статуса из значения поля «Статус ФНС».

    Значение может содержать встроенную дату: ``"LIQUIDATED|2021-06-17"``,
    либо быть псевдокодом: ``"NO_INN"``, ``"PENDING"``.
    Возвращает только код до разделителя.

    Args:
        raw: Сырое значение из SQL.

    Returns:
        Код статуса или ``None``.
    """
    if not raw:
        return None
    return raw.split(_STATUS_DATE_SEP, 1)[0]


# ---------------------------------------------------------------------------
# Стилизация ячеек
# ---------------------------------------------------------------------------


def _style_header(cell: Cell, b: Border) -> None:
    cell.font = Font(name="Calibri", size=11, bold=True, color=_C.HEADER_FG)
    cell.fill = PatternFill(fill_type="solid", fgColor=_C.HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = b  # type: ignore[assignment]


def _style_data(
    cell: Cell,
    header: str,
    value: Any,
    row_idx: int,
    b: Border,
    official_status: str | None,
) -> None:
    """Применяет форматирование к ячейке данных.

    Args:
        cell:            Ячейка openpyxl.
        header:          Заголовок колонки.
        value:           Сырое значение из SQL.
        row_idx:         Индекс строки (для чередования фона).
        b:               Объект Border.
        official_status: Код статуса (только код, без даты) для цвета заливки.
    """
    # --- Статус ФНС: всегда заполнен, поддерживает встроенную дату ----------
    if header == "Статус ФНС":
        raw = str(value) if value is not None else ""
        parts = raw.split(_STATUS_DATE_SEP, 1)
        code = parts[0] if parts[0] else None
        date_part = parts[1] if len(parts) > 1 else None

        label = _STATUS_LABEL.get(code or "", code or "")
        cell.value = f"{label}\n{date_part}" if (label and date_part) else label or None

        bg = _STATUS_COLOR.get(code or "", _C.STATUS_UNKNOWN) if code else _C.ROW_ODD
        cell.fill = PatternFill(fill_type="solid", fgColor=bg)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = b  # type: ignore[assignment]
        return

    # --- Выравнивание -------------------------------------------------------
    if header in _MULTILINE:
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    elif header in _CENTER:
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=False
        )
    else:
        cell.alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=False
        )

    # --- Заливка ------------------------------------------------------------
    if header == "ИНН":
        cell.fill = PatternFill(fill_type="solid", fgColor=_C.INN_BG)
    else:
        cell.fill = _row_fill(row_idx)

    cell.border = b  # type: ignore[assignment]

    # --- Формат и значение --------------------------------------------------
    if header in _TEXT_NUM:
        cell.number_format = "@"
        cell.value = str(value) if value is not None else ""
        return

    if header in _DATE_COLS:
        parsed = _parse_date(value)
        if parsed:
            cell.value = parsed
            cell.number_format = "YYYY-MM-DD"
        else:
            cell.value = value
        return

    cell.value = value


# ---------------------------------------------------------------------------
# Авторасширение и высоты строк
# ---------------------------------------------------------------------------


def _compute_widths(ws: Worksheet, headers: list[str]) -> dict[int, float]:
    """Рассчитывает оптимальные ширины колонок по содержимому."""
    widths: dict[int, float] = {}

    for col_obj in ws.columns:
        max_len = 0
        col_header = ""

        for cell in col_obj:
            assert isinstance(cell, Cell)
            val = cell.value
            if cell.row == 1:
                col_header = str(val) if val is not None else ""
            if val is None:
                continue
            if isinstance(val, datetime.date):
                clen = 10
            elif isinstance(val, str):
                if col_header in _MULTILINE:
                    clen = max((len(ln) for ln in val.split("\n")), default=0)
                else:
                    clen = len(val)
            else:
                clen = len(str(val))
            max_len = max(max_len, clen)

        fc = col_obj[0]
        assert isinstance(fc, Cell) and isinstance(fc.column, int)
        letter = get_column_letter(fc.column)
        width = min(max(max_len + 3, _MIN_W), _MAX_W)
        ws.column_dimensions[letter].width = width
        widths[fc.column] = width

    return widths


def _compute_heights(
    ws: Worksheet,
    headers: list[str],
    col_widths: dict[int, float],
    data_start: int,
    data_end: int,
) -> None:
    """Устанавливает высоту строки по самой «высокой» ячейке.

    Args:
        ws:         Лист openpyxl.
        headers:    Список заголовков.
        col_widths: Словарь {column_index: width}.
        data_start: Первая строка данных.
        data_end:   Последняя строка данных включительно.
    """
    wrap_col_idxs = {i + 1 for i, h in enumerate(headers) if h in _MULTILINE}

    for row in range(data_start, data_end + 1):
        max_lines = 1
        for col in wrap_col_idxs:
            cell = ws.cell(row=row, column=col)
            assert isinstance(cell, Cell)
            if isinstance(cell.value, str) and cell.value:
                lines = _estimate_lines(cell.value, col_widths.get(col, _MIN_W))
                max_lines = max(max_lines, lines)
        ws.row_dimensions[row].height = max(_LINE_H, _LINE_H * max_lines + 4)


# ---------------------------------------------------------------------------
# Главная функция
# ---------------------------------------------------------------------------


def export_sqlite_to_xlsx(db_path: Path | str, xlsx_path: Path | str) -> None:
    """Экспортирует компании из единой БД в XLSX.

    Алгоритм:
      1. Строим SQL из ``config.EXPORT_CONFIG``.
      2. Читаем данные из SQLite.
      3. Проход 1 — заполнение ячеек.
      4. Расчёт ширин колонок.
      5. Проход 2 — высоты строк с учётом финальных ширин.

    Args:
        db_path:   Путь к scraperela.db.
        xlsx_path: Путь к выходному .xlsx.
    """
    db_path = Path(db_path)
    xlsx_path = Path(xlsx_path)

    if not db_path.exists():
        logger.error(f"БД не найдена: {db_path}")
        return

    export_cfg = cfg.EXPORT_CONFIG
    sql, params = _build_export_sql(export_cfg)
    logger.info(f"Фильтры экспорта: {_describe_filters(export_cfg)}")

    conn = sqlite3.connect(db_path)
    try:
        cur = conn.cursor()
        cur.execute(sql, params)
        headers: list[str] = [c[0] for c in cur.description]
        rows: list[tuple[Any, ...]] = cur.fetchall()
    except sqlite3.Error as exc:
        logger.error(f"SQL-ошибка при экспорте: {exc}")
        return
    finally:
        conn.close()

    logger.info(f"Экспорт: {len(rows)} компаний, {len(headers)} колонок.")

    if not rows:
        logger.warning("Нет данных для экспорта — файл не создан.")
        return

    wb = Workbook()
    ws = wb.active
    assert isinstance(ws, Worksheet)
    ws.title = "Реестр ЧОП и ЧОО"
    ws.sheet_view.showGridLines = True
    ws.freeze_panes = "A2"

    b = _make_border()

    for ci, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=header)
        assert isinstance(cell, Cell)
        _style_header(cell, b)
    ws.row_dimensions[1].height = 32

    status_idx = headers.index("Статус ФНС") if "Статус ФНС" in headers else -1

    for ri, row_data in enumerate(rows, 2):
        official_status = _extract_status_code(
            str(row_data[status_idx])
            if status_idx >= 0 and row_data[status_idx]
            else None
        )
        for ci, value in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci)
            assert isinstance(cell, Cell)
            _style_data(
                cell=cell,
                header=headers[ci - 1],
                value=value,
                row_idx=ri,
                b=b,
                official_status=official_status,
            )

    col_widths = _compute_widths(ws, headers)
    _compute_heights(ws, headers, col_widths, 2, len(rows) + 1)

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    logger.info(f"XLSX сохранён: {xlsx_path}")
